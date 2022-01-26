#! python3
# updateproduce.py - corrects costs in produce sales spreadsheet

import openpyxl

wb = openpyxl.load_workbook('chapter13/produceSales.xlsx')
sheet = wb['Sheet']

PRICE_UPDATES = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
}

for row_num in range(2, sheet.max_row):
    produce_name = sheet.cell(row=row_num, column=1).value
    if produce_name in PRICE_UPDATES:
        sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]

wb.save('chapter13/updatedProduceSales.xlsx')