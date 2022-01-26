import openpyxl, sys

def make_table(n):
    # create new spreadsheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    bold = openpyxl.styles.Font(bold=True)

    # table headers
    for i in range(1,n+1):
        sheet.cell(1, i+1).value = i
        sheet.cell(1, i+1).font = bold
        sheet.cell(i+1, 1).value = i
        sheet.cell(i+1, 1).font = bold

    # iterate through inside cells and get products
    #for row in sheet.iter_rows(min_row=2, max_row=n+1, max_col=n+1):
    for i in range(1, n+1):
        for j in range(1, n+1):
            sheet.cell(i+1, j+1).value = i * j
    
    wb.save('chapter13/multiplication_results.xlsx')
    

def main():
    if len(sys.argv) < 2:
        print('please pass a number N for table dimensions')
    else:
        make_table(int(sys.argv[1]))


if __name__ == '__main__':
    main()