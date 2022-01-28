import openpyxl, csv, os

for excel_fn in os.listdir('.'):
    if not excel_fn.endswith('.xlsx'):
        continue
    with open(excel_fn, 'rb') as ef:
        wb = openpyxl.load_workbook(ef)
        for sn in wb.sheetnames:
            sheet = wb[sn]
            csv_fn = f'{excel_fn[:-5]}_{sn}.csv'
            with open(csv_fn, 'w', newline='') as cf:
                csv_writer = csv.writer(cf)
                for row_num in range(1, sheet.max_row + 1):
                    row_data = [sheet.cell(row=row_num, column=i).value for i in range(1, sheet.max_column + 1)]
                    csv_writer.writerow(row_data)
